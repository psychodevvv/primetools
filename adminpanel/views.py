import io
import openpyxl
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, JsonResponse
from django.views.decorators.http import require_POST
from django.contrib import messages
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Q
from django.utils.text import slugify
from shop.models import Category, Product, Order, VideoReview
from accounts.models import Customer
from .models import AdminUser


def _paginate(request, queryset, per_page):
    """Возвращает (page_obj, page_range, base_qs) для постраничного вывода."""
    paginator = Paginator(queryset, per_page)
    page_obj = paginator.get_page(request.GET.get('page'))
    page_range = paginator.get_elided_page_range(
        page_obj.number, on_each_side=1, on_ends=1)
    params = request.GET.copy()
    params.pop('page', None)
    return page_obj, page_range, params.urlencode()


def admin_required(view_func):
    def wrapper(request, *args, **kwargs):
        if not request.session.get('admin_id'):
            return redirect('admin_login')
        return view_func(request, *args, **kwargs)
    wrapper.__name__ = view_func.__name__
    return wrapper


def admin_login(request):
    if request.session.get('admin_id'):
        return redirect('admin_dashboard')

    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '')
        try:
            admin = AdminUser.objects.get(username=username)
            if admin.check_password(password):
                request.session['admin_id'] = admin.pk
                request.session['admin_username'] = admin.username
                return redirect('admin_dashboard')
            else:
                messages.error(request, 'Неверный пароль.')
        except AdminUser.DoesNotExist:
            messages.error(request, 'Пользователь не найден.')

    return render(request, 'adminpanel/login.html')


def admin_logout(request):
    request.session.pop('admin_id', None)
    request.session.pop('admin_username', None)
    return redirect('admin_login')


@admin_required
def dashboard(request):
    orders = Order.objects.prefetch_related('items').order_by('-created_at')[:20]
    total_orders = Order.objects.count()
    new_orders = Order.objects.filter(status='new').count()
    total_products = Product.objects.count()
    total_categories = Category.objects.count()
    return render(request, 'adminpanel/dashboard.html', {
        'orders': orders,
        'total_orders': total_orders,
        'new_orders': new_orders,
        'total_products': total_products,
        'total_categories': total_categories,
    })


@admin_required
def orders_list(request):
    status_filter = request.GET.get('status', '')
    search_query = request.GET.get('q', '').strip()
    orders = Order.objects.prefetch_related('items').order_by('-created_at')
    if status_filter:
        orders = orders.filter(status=status_filter)
    if search_query:
        cond = (Q(customer_name__icontains=search_query) |
                Q(customer_phone__icontains=search_query))
        if search_query.isdigit():
            cond |= Q(pk=search_query)
        orders = orders.filter(cond)
    page_obj, page_range, base_qs = _paginate(request, orders, 30)
    return render(request, 'adminpanel/orders.html', {
        'orders': page_obj,
        'page_obj': page_obj,
        'page_range': page_range,
        'base_qs': base_qs,
        'status_filter': status_filter,
        'search_query': search_query,
    })


@admin_required
def order_detail(request, pk):
    order = get_object_or_404(Order, pk=pk)
    if request.method == 'POST':
        new_status = request.POST.get('status')
        if new_status in dict(Order.STATUS_CHOICES):
            order.status = new_status
            order.save()
            messages.success(request, 'Статус обновлён.')
        return redirect('admin_order_detail', pk=pk)
    return render(request, 'adminpanel/order_detail.html', {'order': order})


@admin_required
def products_list(request):
    cat_filter = request.GET.get('cat', '')
    search_query = request.GET.get('q', '').strip()
    products = Product.objects.select_related('category').order_by('category__name', 'name')
    if cat_filter:
        products = products.filter(category__slug=cat_filter)
    if search_query:
        products = products.filter(
            Q(name__icontains=search_query) |
            Q(article__icontains=search_query) |
            Q(brand__icontains=search_query)
        )
    page_obj, page_range, base_qs = _paginate(request, products, 50)
    categories = Category.objects.all()
    return render(request, 'adminpanel/products.html', {
        'products': page_obj,
        'page_obj': page_obj,
        'page_range': page_range,
        'base_qs': base_qs,
        'total_count': page_obj.paginator.count,
        'categories': categories,
        'cat_filter': cat_filter,
        'search_query': search_query,
    })


@admin_required
def product_edit(request, pk=None):
    product = get_object_or_404(Product, pk=pk) if pk else None
    categories = Category.objects.all()

    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        brand = request.POST.get('brand', '').strip()
        article = request.POST.get('article', '').strip()
        description = request.POST.get('description', '').strip()
        price = request.POST.get('price', '0')
        image_url = request.POST.get('image_url', '').strip()
        category_id = request.POST.get('category')
        in_stock = request.POST.get('in_stock') == 'on'

        if not name or not price or not category_id:
            messages.error(request, 'Заполните обязательные поля.')
        else:
            cat = get_object_or_404(Category, pk=category_id)
            if product:
                product.name = name
                product.brand = brand
                product.article = article
                product.description = description
                product.price = price
                product.image_url = image_url
                product.category = cat
                product.in_stock = in_stock
                product.save()
                messages.success(request, 'Товар обновлён.')
            else:
                Product.objects.create(
                    name=name, brand=brand, article=article,
                    description=description, price=price,
                    image_url=image_url, category=cat, in_stock=in_stock,
                )
                messages.success(request, 'Товар добавлен.')
            return redirect('admin_products')

    return render(request, 'adminpanel/product_edit.html', {
        'product': product,
        'categories': categories,
    })


@admin_required
@require_POST
def product_delete(request, pk):
    product = get_object_or_404(Product, pk=pk)
    product.delete()
    messages.success(request, 'Товар удалён.')
    return redirect('admin_products')


@admin_required
def categories_list(request):
    categories = Category.objects.select_related('parent').order_by('parent__name', 'name')
    page_obj, page_range, base_qs = _paginate(request, categories, 30)
    return render(request, 'adminpanel/categories.html', {
        'categories': page_obj,
        'page_obj': page_obj,
        'page_range': page_range,
        'base_qs': base_qs,
    })


@admin_required
def category_edit(request, pk=None):
    category = get_object_or_404(Category, pk=pk) if pk else None

    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        slug = request.POST.get('slug', '').strip() or slugify(name, allow_unicode=True)
        description = request.POST.get('description', '').strip()
        image_url = request.POST.get('image_url', '').strip()
        order = request.POST.get('order') or 0
        parent_id = request.POST.get('parent') or None
        parent = Category.objects.filter(pk=parent_id).first() if parent_id else None
        if parent and category and parent.pk == category.pk:
            parent = None  # категория не может быть родителем самой себе

        if category:
            category.name = name
            category.slug = slug
            category.description = description
            category.image_url = image_url
            category.order = order
            category.parent = parent
            category.save()
            messages.success(request, 'Категория обновлена.')
        else:
            Category.objects.create(
                name=name, slug=slug, description=description,
                image_url=image_url, order=order, parent=parent,
            )
            messages.success(request, 'Категория создана.')
        return redirect('admin_categories')

    all_categories = Category.objects.order_by('name')
    if category:
        all_categories = all_categories.exclude(pk=category.pk)
    return render(request, 'adminpanel/category_edit.html', {
        'category': category,
        'all_categories': all_categories,
    })


@admin_required
def excel_download_template(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Товары'

    headers = ['Название*', 'Категория (slug)*', 'Бренд', 'Артикул', 'Цена*', 'Описание', 'Ссылка на фото', 'В наличии (да/нет)']
    ws.append(headers)

    for col in ws.iter_cols(min_row=1, max_row=1):
        for cell in col:
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill(fill_type='solid', fgColor='FF6600')

    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 40
    ws.column_dimensions['G'].width = 40
    ws.column_dimensions['H'].width = 15

    cats = Category.objects.values_list('slug', 'name')
    notes_ws = wb.create_sheet('Категории')
    notes_ws.append(['slug', 'Название'])
    for slug, name in cats:
        notes_ws.append([slug, name])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="products_template.xlsx"'
    return response


# ─── Гибкий импорт прайс-листов ───────────────────────────────────────────
# Подсказки для распознавания столбцов. Формат: (ключевое слово, вес).
# Чем выше вес — тем приоритетнее столбец при совпадении.
COLUMN_HINTS = {
    'name':     [('наименование', 3), ('название', 3), ('номенклатура', 2),
                 ('товар', 2), ('продукт', 2), ('наимен', 2), ('name', 2)],
    'brand':    [('бренд', 3), ('производитель', 2), ('изготовитель', 2),
                 ('марка', 2), ('brand', 2)],
    'article':  [('артикул', 3), ('sku', 3), ('код товара', 3),
                 ('арт.', 2), ('артик', 2), ('код', 2), ('article', 2)],
    'category': [('подкатегория', 3), ('категория', 3), ('категории', 3),
                 ('раздел', 2), ('группа товаров', 2), ('группа', 2), ('category', 2)],
    'description': [('описание', 3), ('характеристик', 2), ('description', 2),
                    ('примечание', 1)],
    'image':    [('ссылка на фото', 4), ('url фото', 4), ('фото', 3),
                 ('изображен', 3), ('картинк', 3), ('image', 2), ('img', 2)],
    'source':   [('ссылка на товар', 3), ('ссылка на страниц', 3),
                 ('страница товара', 3), ('ссылка', 2), ('url товара', 2),
                 ('product url', 2), ('url', 1), ('link', 1)],
    'price':    [('рекоменд', 5), ('розничн', 5), ('цена продаж', 5),
                 ('рек. цена', 5), ('рек.цена', 5), ('цена', 3),
                 ('стоимост', 3), ('price', 3), ('оптов', 2), ('закуп', 1)],
    'stock':    [('в наличии', 3), ('наличие', 2)],
}


def _norm(value):
    """Нормализует текст заголовка для сравнения."""
    return str(value or '').strip().lower().replace('ё', 'е')


def _detect_header_row(ws, max_scan=15):
    """Находит строку заголовков — ту, где больше всего распознанных столбцов."""
    best_row, best_score = 1, -1
    for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True), start=1):
        cells = [_norm(c) for c in row if c not in (None, '')]
        if not cells:
            continue
        score = 0
        for hints in COLUMN_HINTS.values():
            if any(kw in cell for kw, _ in hints for cell in cells):
                score += 1
        if score > best_score:
            best_score, best_row = score, idx
    return best_row if best_score >= 2 else 1


def _map_columns(header_cells):
    """Сопоставляет столбцы файла с полями товара по заголовкам."""
    single = {}          # поле -> (индекс, вес)
    price_cols = []      # [(индекс, вес)] — все столбцы-кандидаты с ценой
    for col_idx, cell in enumerate(header_cells):
        h = _norm(cell)
        if not h:
            continue
        for field, hints in COLUMN_HINTS.items():
            best = 0
            for kw, weight in hints:
                if kw in h and weight > best:
                    best = weight
            if not best:
                continue
            if field == 'price':
                price_cols.append((col_idx, best))
            elif field not in single or single[field][1] < best:
                single[field] = (col_idx, best)
    mapping = {f: idx for f, (idx, _) in single.items()}
    mapping['price'] = [idx for idx, _ in sorted(price_cols, key=lambda x: -x[1])]
    return mapping


def _cell(row, idx):
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def _parse_price(value):
    if value is None:
        return None
    try:
        cleaned = str(value).replace('\xa0', '').replace(' ', '').replace(',', '.')
        cleaned = ''.join(c for c in cleaned if c.isdigit() or c == '.')
        return float(cleaned) if cleaned else None
    except (ValueError, TypeError):
        return None


# ─── Импорт каталога в формате YML (Yandex Market) ────────────────────────
import html as _htmllib
import re as _re


def _yml_html_to_text(raw):
    """Преобразует HTML-описание из YML-фида в читаемый текст с переносами."""
    if not raw:
        return ''
    t = _re.sub(r'<li[^>]*>', '\n• ', raw, flags=_re.IGNORECASE)
    t = _re.sub(r'<br\s*/?>', '\n', t, flags=_re.IGNORECASE)
    t = _re.sub(r'</(?:p|h[1-6]|ul|ol|div|li|tr)\s*>', '\n', t, flags=_re.IGNORECASE)
    t = _re.sub(r'<h[1-6][^>]*>', '\n', t, flags=_re.IGNORECASE)
    t = _re.sub(r'<[^>]+>', '', t)
    t = _htmllib.unescape(t)
    lines = [_re.sub(r'[ \t]+', ' ', ln).strip() for ln in t.splitlines()]
    return '\n'.join(ln for ln in lines if ln)


def _yml_offer_dict(offer):
    """Извлекает данные товара из элемента <offer>."""
    def txt(tag):
        el = offer.find(tag)
        return (el.text or '').strip() if el is not None and el.text else ''

    name = txt('name')
    if not name:
        return None

    price = None
    for tag in ('price_rrc', 'price'):
        try:
            p = float((txt(tag) or '0').replace(',', '.'))
        except ValueError:
            p = 0
        if p > 0:
            price = p
            break
    if price is None:
        return None

    pics = []  # все фото товара
    for pic in offer.findall('picture'):
        u = (pic.text or '').strip()
        if u.lower().startswith(('http://', 'https://')) and u not in pics:
            pics.append(u)
    image_url = pics[0] if pics else ''
    gallery = '\n'.join(pics)

    source_url = txt('url')
    if not source_url.lower().startswith(('http://', 'https://')):
        source_url = ''

    params = []
    for p in offer.findall('param'):
        pn = (p.get('name') or '').strip()
        pv = (p.text or '').strip()
        if pn:
            params.append(f'{pn}: {pv}' if pv else pn)

    return dict(
        name=name[:500],
        brand=txt('vendor')[:200],
        article=txt('vendorCode')[:200],
        price=price,
        cat_yml_id=txt('categoryId'),
        image_url=image_url,
        gallery=gallery,
        source_url=source_url,
        description=_yml_html_to_text(txt('description'))[:6000],
        characteristics='\n'.join(params)[:4000],
        in_stock=(offer.get('available') or 'true').lower() != 'false',
    )


def _import_yml(request, upload):
    """Импортирует каталог из YML-фида: товары + дерево категорий (parentId)."""
    import xml.etree.ElementTree as ET

    try:
        upload.seek(0)
        yml_cats, offers = {}, []
        for _event, elem in ET.iterparse(upload, events=('end',)):
            if elem.tag == 'category':
                cid = elem.get('id')
                if cid:
                    yml_cats[cid] = {
                        'name': (elem.text or '').strip() or 'Без названия',
                        'parent': elem.get('parentId'),
                    }
                elem.clear()
            elif elem.tag == 'offer':
                row = _yml_offer_dict(elem)
                if row:
                    offers.append(row)
                elem.clear()
    except Exception as e:
        messages.error(request, f'Не удалось прочитать YML-файл: {e}')
        return render(request, 'adminpanel/excel_import.html')

    if not offers:
        messages.warning(request, 'В YML-файле не найдено товаров.')
        return render(request, 'adminpanel/excel_import.html')

    # Создаём дерево категорий с подкатегориями (по name + parent — идемпотентно).
    resolved = {}

    def resolve_yml_cat(yid, depth=0):
        if yid in resolved:
            return resolved[yid]
        info = yml_cats.get(yid)
        if not info or depth > 12:
            return None
        parent = resolve_yml_cat(info['parent'], depth + 1) if info['parent'] else None
        cat = Category.objects.filter(name=info['name'][:200], parent=parent).first()
        if cat is None:
            cat = _make_category(info['name'])
            if parent:
                cat.parent = parent
                cat.save(update_fields=['parent'])
        resolved[yid] = cat
        return cat

    default_cat = [None]

    def get_category(yid):
        cat = resolve_yml_cat(yid) if yid else None
        if cat is None:
            if default_cat[0] is None:
                default_cat[0] = (
                    Category.objects.filter(name=DEFAULT_CATEGORY_NAME, parent=None).first()
                    or _make_category(DEFAULT_CATEGORY_NAME))
            cat = default_cat[0]
        return cat

    existing = {p.name: p for p in Product.objects.all()}
    to_create, to_update = {}, {}
    for row in offers:
        category = get_category(row.pop('cat_yml_id'))
        name = row.pop('name')
        fields = dict(row, category=category)
        if name in existing:
            product = existing[name]
            for k, v in fields.items():
                setattr(product, k, v)
            to_update[name] = product
        elif name in to_create:
            for k, v in fields.items():
                setattr(to_create[name], k, v)
        else:
            to_create[name] = Product(name=name, **fields)

    with transaction.atomic():
        if to_create:
            Product.objects.bulk_create(list(to_create.values()), batch_size=500)
        if to_update:
            Product.objects.bulk_update(
                list(to_update.values()),
                ['category', 'brand', 'article', 'price', 'description',
                 'characteristics', 'image_url', 'gallery', 'source_url', 'in_stock'],
                batch_size=500,
            )

    messages.success(
        request,
        f'YML-импорт завершён: создано {len(to_create)}, обновлено {len(to_update)}. '
        f'Цена взята из «price_rrc» (рекомендованная розничная).'
    )
    return render(request, 'adminpanel/excel_import.html')


@admin_required
def excel_import(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        if excel_file.name.lower().endswith(('.yml', '.xml')):
            return _import_yml(request, excel_file)
        try:
            wb = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)
            ws = wb.active

            header_row_idx = _detect_header_row(ws)
            header_cells, rows = [], []
            for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
                if i < header_row_idx:
                    continue
                if i == header_row_idx:
                    header_cells = list(row)
                else:
                    rows.append(row)
            wb.close()

            cols = _map_columns(header_cells)
            # «Ссылка на фото» не должна одновременно считаться ссылкой-источником
            if cols.get('source') is not None and cols.get('source') == cols.get('image'):
                cols.pop('source')
            if 'name' not in cols:
                messages.error(
                    request,
                    'Не удалось определить столбец с названием товара. '
                    'Проверьте, что в файле есть заголовок «Наименование» / «Название».'
                )
                return render(request, 'adminpanel/excel_import.html')

            # Кеш категорий: по slug и по нормализованному названию
            cat_cache = {}
            for c in Category.objects.all():
                cat_cache[c.slug] = c
                cat_cache['name:' + _norm(c.name)] = c

            def resolve_category(raw):
                """Находит категорию по slug/названию или создаёт новую."""
                raw = str(raw or '').strip()
                if not raw:
                    return None
                key_slug = slugify(raw, allow_unicode=False)
                if raw in cat_cache:
                    return cat_cache[raw]
                if key_slug and key_slug in cat_cache:
                    return cat_cache[key_slug]
                if 'name:' + _norm(raw) in cat_cache:
                    return cat_cache['name:' + _norm(raw)]
                slug = key_slug or ('cat-' + str(len(cat_cache) + 1))
                base, n = slug, 2
                while Category.objects.filter(slug=slug).exists():
                    slug = f'{base}-{n}'; n += 1
                cat = Category.objects.create(name=raw[:200], slug=slug)
                cat_cache[raw] = cat
                cat_cache[slug] = cat
                cat_cache['name:' + _norm(raw)] = cat
                return cat

            default_category = None
            if 'category' not in cols:
                default_category = resolve_category('Импортированные товары')

            existing = {p.name: p for p in Product.objects.all()}
            to_create, to_update = {}, {}
            created = updated = skipped = 0
            errors = []

            for offset, row in enumerate(rows, start=header_row_idx + 1):
                name = str(_cell(row, cols['name']) or '').strip()
                if not name:
                    continue

                price = None
                for pcol in cols.get('price', []):
                    price = _parse_price(_cell(row, pcol))
                    if price is not None:
                        break
                if price is None:
                    skipped += 1
                    if len(errors) < 8:
                        errors.append(f'Строка {offset}: не найдена цена — пропущено.')
                    continue

                if default_category is not None:
                    category = default_category
                else:
                    category = resolve_category(_cell(row, cols['category']))
                    if category is None:
                        category = default_category = resolve_category('Импортированные товары')

                brand = str(_cell(row, cols.get('brand')) or '').strip()[:200]
                article = str(_cell(row, cols.get('article')) or '').strip()[:200]
                description = str(_cell(row, cols.get('description')) or '').strip()
                image_url = str(_cell(row, cols.get('image')) or '').strip()
                if not image_url.lower().startswith(('http://', 'https://')):
                    image_url = ''
                source_url = str(_cell(row, cols.get('source')) or '').strip()
                if not source_url.lower().startswith(('http://', 'https://')):
                    source_url = ''

                in_stock = True
                if 'stock' in cols:
                    sv = _cell(row, cols['stock'])
                    if sv is not None and str(sv).strip() != '':
                        sl = _norm(sv)
                        if sl in ('нет', 'no', 'false', '0', '-', 'отсутствует'):
                            in_stock = False
                        elif isinstance(sv, (int, float)):
                            in_stock = sv > 0

                fields = dict(
                    category=category, brand=brand, article=article,
                    price=price, description=description,
                    image_url=image_url, source_url=source_url, in_stock=in_stock,
                )

                if name in existing:
                    p = existing[name]
                    for k, v in fields.items():
                        setattr(p, k, v)
                    to_update[name] = p
                elif name in to_create:
                    p = to_create[name]
                    for k, v in fields.items():
                        setattr(p, k, v)
                else:
                    to_create[name] = Product(name=name[:500], **fields)

            with transaction.atomic():
                if to_create:
                    Product.objects.bulk_create(list(to_create.values()), batch_size=500)
                    created = len(to_create)
                if to_update:
                    Product.objects.bulk_update(
                        list(to_update.values()),
                        ['category', 'brand', 'article', 'price',
                         'description', 'image_url', 'source_url', 'in_stock'],
                        batch_size=500,
                    )
                    updated = len(to_update)

            detected = ', '.join(
                label for key, label in [
                    ('name', 'название'), ('price', 'цена'), ('brand', 'бренд'),
                    ('article', 'артикул'), ('category', 'категория'),
                    ('description', 'описание'), ('image', 'фото'),
                    ('source', 'ссылка-источник'), ('stock', 'наличие'),
                ] if (cols.get(key) not in (None, []))
            )
            msg = (f'Импорт завершён: создано {created}, обновлено {updated}.'
                   f' Распознаны столбцы: {detected}.')
            if skipped:
                msg += f' Пропущено строк без цены: {skipped}.'
            messages.success(request, msg)
            for err in errors[:5]:
                messages.warning(request, err)

            # Если указан сайт поставщика — ищем фото/описания по артикулу.
            site_url = (request.POST.get('site_url') or '').strip()
            if site_url:
                done, found, remaining = _run_site_search(site_url, time_budget=55)
                if found:
                    s = (f'С сайта {site_url}: проверено {done} товаров, '
                         f'найдено и загружено фото для {found}.')
                    if remaining > 0:
                        s += (f' Осталось {remaining} — продолжите в разделе '
                               f'«Обработка товаров».')
                    messages.success(request, s)
                else:
                    messages.warning(
                        request,
                        f'На сайте {site_url} не удалось найти товары по артикулу — '
                        f'проверьте адрес сайта.')

        except Exception as e:
            messages.error(request, f'Не удалось обработать файл: {e}')

    return render(request, 'adminpanel/excel_import.html')


# ─── Категории и загрузка фото/описаний ───────────────────────────────────
DEFAULT_CATEGORY_NAME = 'Импортированные товары'
INFO_FETCH_CHUNK = 120


def _make_category(name):
    """Создаёт категорию с уникальным slug."""
    base = slugify(name, allow_unicode=True) or 'cat'
    slug, n = base, 2
    while Category.objects.filter(slug=slug).exists():
        slug = f'{base}-{n}'
        n += 1
    return Category.objects.create(name=name[:200], slug=slug)


@admin_required
def ai_dashboard(request):
    default_cats = Category.objects.filter(name=DEFAULT_CATEGORY_NAME)
    uncategorized = Product.objects.filter(category__in=default_cats).count()
    need_info = (
        Product.objects.exclude(source_url='')
        .filter(Q(image_url='') | Q(description='') | Q(characteristics='')).count()
    )
    no_link = (
        Product.objects.filter(image_url='', source_url='')
        .exclude(article='').count()
    )
    return render(request, 'adminpanel/ai.html', {
        'uncategorized': uncategorized,
        'need_info': need_info,
        'no_link': no_link,
    })


@admin_required
@require_POST
def ai_categorize(request):
    """Распознаёт категории по ключевым словам — мгновенно, для всех товаров."""
    from shop.ai import classify

    default_cats = list(Category.objects.filter(name=DEFAULT_CATEGORY_NAME))
    products = list(Product.objects.filter(category__in=default_cats))
    if not products:
        messages.success(request, 'Все товары уже распределены по категориям.')
        return redirect('admin_ai')

    cat_by_name = {c.name.lower(): c for c in Category.objects.all()}
    to_update, moved, new_cats, unmatched = [], 0, 0, 0
    for product in products:
        name = classify(product.name)
        if not name:
            unmatched += 1
            continue
        category = cat_by_name.get(name.lower())
        if category is None:
            category = _make_category(name)
            cat_by_name[name.lower()] = category
            new_cats += 1
        if product.category_id != category.pk:
            product.category = category
            to_update.append(product)
            moved += 1
    if to_update:
        Product.objects.bulk_update(to_update, ['category'], batch_size=500)

    msg = f'Распределено по категориям: {moved} товаров.'
    if new_cats:
        msg += f' Создано категорий: {new_cats}.'
    if unmatched:
        msg += (f' Не удалось определить: {unmatched} '
                f'(остались в «{DEFAULT_CATEGORY_NAME}»).')
    messages.success(request, msg)
    return redirect('admin_ai')


@admin_required
@require_POST
def ai_fetch_images(request):
    """Загружает фото и описания со страниц товаров (партиями за один клик)."""
    import time
    from concurrent.futures import ThreadPoolExecutor
    from shop.images import fetch_product_info

    # Снимок id — за один клик каждый товар обрабатывается ровно один раз.
    pks = list(
        Product.objects.exclude(source_url='')
        .filter(Q(image_url='') | Q(description='') | Q(characteristics=''))
        .values_list('pk', flat=True)
    )
    if not pks:
        messages.success(request, 'Нет товаров, для которых нужно подобрать фото/описание.')
        return redirect('admin_ai')

    deadline = time.time() + 45
    idx = total_done = got_image = got_desc = got_char = 0

    while idx < len(pks) and time.time() < deadline:
        chunk = pks[idx:idx + INFO_FETCH_CHUNK]
        idx += len(chunk)
        batch = list(Product.objects.filter(pk__in=chunk))

        # Страницы качаем параллельно, ORM трогаем только в основном потоке.
        info = {}
        with ThreadPoolExecutor(max_workers=30) as pool:
            for pk, data in pool.map(
                lambda p: (p.pk, fetch_product_info(p.source_url)), batch
            ):
                info[pk] = data

        to_update = []
        for product in batch:
            data = info.get(product.pk) or {}
            changed = False
            if not product.image_url and data.get('image'):
                product.image_url = data['image']
                got_image += 1
                changed = True
            if not product.description and data.get('description'):
                product.description = data['description']
                got_desc += 1
                changed = True
            if not product.characteristics and data.get('characteristics'):
                product.characteristics = data['characteristics']
                got_char += 1
                changed = True
            if changed:
                to_update.append(product)
        if to_update:
            Product.objects.bulk_update(
                to_update, ['image_url', 'description', 'characteristics'])
        total_done += len(batch)

    remaining = len(pks) - total_done
    msg = (f'Обработано {total_done} товаров — фото: {got_image}, '
           f'описаний: {got_desc}, характеристик: {got_char}.')
    if remaining > 0:
        msg += f' Осталось ещё {remaining} — нажмите кнопку повторно.'
    messages.success(request, msg)
    return redirect('admin_ai')


def _run_site_search(site, time_budget=50):
    """Ищет товары без фото по артикулу на сайте поставщика и забирает данные.

    Возвращает (проверено, найдено_на_сайте, осталось).
    """
    import time
    from concurrent.futures import ThreadPoolExecutor
    from shop.images import search_product_on_site, fetch_product_info

    pks = list(
        Product.objects.filter(image_url='', source_url='')
        .exclude(article='').values_list('pk', flat=True)
    )
    if not pks:
        return 0, 0, 0

    deadline = time.time() + time_budget
    pattern = [None]            # рабочий шаблон поиска (кешируется)
    idx = total_done = found_url = 0
    CHUNK = 40

    while idx < len(pks) and time.time() < deadline:
        chunk = pks[idx:idx + CHUNK]
        idx += len(chunk)
        batch = list(Product.objects.filter(pk__in=chunk))

        located = {}
        with ThreadPoolExecutor(max_workers=12) as pool:
            for pk, url, pat in pool.map(
                lambda p: (p.pk,) + search_product_on_site(site, p.article, pattern[0]),
                batch,
            ):
                located[pk] = url
                if pat and pattern[0] is None:
                    pattern[0] = pat

        urls = {pk: u for pk, u in located.items() if u}
        info = {}
        if urls:
            with ThreadPoolExecutor(max_workers=20) as pool:
                for pk, data in pool.map(
                    lambda kv: (kv[0], fetch_product_info(kv[1])), list(urls.items())
                ):
                    info[pk] = data

        to_update = []
        for product in batch:
            url = located.get(product.pk)
            if not url:
                continue
            product.source_url = url
            found_url += 1
            data = info.get(product.pk) or {}
            if data.get('image'):
                product.image_url = data['image']
            if data.get('description') and not product.description:
                product.description = data['description']
            if data.get('characteristics') and not product.characteristics:
                product.characteristics = data['characteristics']
            to_update.append(product)
        if to_update:
            Product.objects.bulk_update(
                to_update,
                ['source_url', 'image_url', 'description', 'characteristics'])
        total_done += len(batch)

    return total_done, found_url, len(pks) - total_done


@admin_required
@require_POST
def ai_site_search(request):
    """Поиск товаров по артикулу на сайте поставщика (раздел «Обработка»)."""
    site = (request.POST.get('site_url') or '').strip()
    if not site:
        messages.error(request, 'Укажите ссылку на сайт поставщика.')
        return redirect('admin_ai')

    done, found, remaining = _run_site_search(site)
    if done == 0:
        messages.success(request, 'Нет товаров без фото и ссылки (с артикулом).')
    elif found == 0:
        messages.warning(
            request,
            'На сайте не удалось найти товары по артикулу. '
            'Проверьте ссылку на сайт или используйте YML-фид.'
        )
    else:
        msg = f'Проверено {done} товаров — найдено на сайте: {found}.'
        if remaining > 0:
            msg += f' Осталось ещё {remaining} — нажмите кнопку повторно.'
        messages.success(request, msg)
    return redirect('admin_ai')


# ─── Покупатели ───────────────────────────────────────────────────────────
@admin_required
def customers_list(request):
    search_query = request.GET.get('q', '').strip()
    customers = Customer.objects.order_by('-created_at')
    if search_query:
        cond = (Q(first_name__icontains=search_query) |
                Q(last_name__icontains=search_query) |
                Q(phone__icontains=search_query))
        if search_query.isdigit():
            cond |= Q(pk=search_query)
        customers = customers.filter(cond)
    page_obj, page_range, base_qs = _paginate(request, customers, 40)
    return render(request, 'adminpanel/customers.html', {
        'customers': page_obj,
        'page_obj': page_obj,
        'page_range': page_range,
        'base_qs': base_qs,
        'search_query': search_query,
    })


def _norm_admin_phone(raw):
    digits = ''.join(c for c in (raw or '') if c.isdigit())
    if digits.startswith('8'):
        digits = '7' + digits[1:]
    if len(digits) == 10:
        digits = '7' + digits
    return digits if len(digits) == 11 and digits.startswith('7') else ''


@admin_required
def customer_edit(request, pk):
    customer = get_object_or_404(Customer, pk=pk)
    if request.method == 'POST':
        first_name = request.POST.get('first_name', '').strip()
        last_name = request.POST.get('last_name', '').strip()
        phone = _norm_admin_phone(request.POST.get('phone', ''))
        if not first_name or not phone:
            messages.error(request, 'Укажите имя и корректный номер телефона.')
        elif Customer.objects.filter(phone=phone).exclude(pk=customer.pk).exists():
            messages.error(request, 'Этот номер уже занят другим покупателем.')
        else:
            customer.first_name = first_name
            customer.last_name = last_name
            customer.phone = phone
            customer.save()
            messages.success(request, 'Данные покупателя обновлены.')
            return redirect('admin_customers')
    return render(request, 'adminpanel/customer_edit.html', {'customer': customer})


@admin_required
@require_POST
def customer_delete(request, pk):
    customer = get_object_or_404(Customer, pk=pk)
    customer.delete()
    messages.success(request, 'Покупатель удалён из базы.')
    return redirect('admin_customers')


# ─── Видео-обзоры на главной ─────────────────────────────────────────────
@admin_required
def videoreviews_list(request):
    videos = VideoReview.objects.all()
    return render(request, 'adminpanel/videos.html', {'videos': videos})


@admin_required
def videoreview_edit(request, pk=None):
    video = get_object_or_404(VideoReview, pk=pk) if pk else None
    if request.method == 'POST':
        title = request.POST.get('title', '').strip()
        video_url = request.POST.get('video_url', '').strip()
        thumbnail_url = request.POST.get('thumbnail_url', '').strip()
        try:
            order = int(request.POST.get('order', 0))
        except ValueError:
            order = 0
        is_active = request.POST.get('is_active') == '1'

        if video is None:
            video = VideoReview()
        video.title = title[:200]
        video.video_url = video_url
        video.thumbnail_url = thumbnail_url
        video.order = order
        video.is_active = is_active
        if request.FILES.get('video_file'):
            video.video_file = request.FILES['video_file']
        elif request.POST.get('clear_video_file') == '1' and video.video_file:
            video.video_file.delete(save=False)
            video.video_file = None
        if not video.video_url and not video.video_file:
            messages.error(request, 'Укажите ссылку или загрузите видео-файл.')
            return render(request, 'adminpanel/video_edit.html', {'video': video})
        video.save()
        messages.success(request, 'Видео-обзор сохранён.')
        return redirect('admin_videos')
    return render(request, 'adminpanel/video_edit.html', {'video': video})


@admin_required
@require_POST
def videoreview_delete(request, pk):
    video = get_object_or_404(VideoReview, pk=pk)
    if video.video_file:
        video.video_file.delete(save=False)
    video.delete()
    messages.success(request, 'Видео-обзор удалён.')
    return redirect('admin_videos')
