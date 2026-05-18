"""Полный импорт каталога: очищает базу и загружает прайс-листы поставщиков.

Запуск:  python manage.py import_catalog

Файлы берутся из папки «Загрузки» пользователя. Категории определяются
автоматически — для прайсов с инструментом по ключевым словам названия,
для YML-фида по дереву категорий, для Al-Style по разделам прайса.
Категория «Импортированные товары» не используется.
"""
import html
import os
import re
import xml.etree.ElementTree as ET

import openpyxl
from django.core.management.base import BaseCommand
from django.db import transaction
from django.utils.text import slugify

from shop.ai import classify
from shop.models import Category, OrderItem, Product

DOWNLOADS = os.path.join(os.path.expanduser('~'), 'Downloads')
MISC_CATEGORY = 'Прочий инструмент и оборудование'
ELECTRONICS_PARENT = 'Электроника и гаджеты'

_TAG_RE = re.compile(r'<[^>]+>')
_WS_RE = re.compile(r'[ \t]+')


def strip_html(text):
    """Превращает HTML-описание в обычный текст."""
    if not text:
        return ''
    text = _TAG_RE.sub('\n', str(text))
    text = html.unescape(text)
    lines = [_WS_RE.sub(' ', ln).strip() for ln in text.splitlines()]
    return '\n'.join(ln for ln in lines if ln).strip()


def parse_price(value):
    """Парсит цену из ячейки. Возвращает float или None."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value) if value > 0 else None
    s = str(value).strip().replace('\xa0', '').replace(' ', '').replace(',', '.')
    s = re.sub(r'[^0-9.]', '', s)
    if not s:
        return None
    try:
        p = float(s)
    except ValueError:
        return None
    return p if p > 0 else None


def as_stock(value):
    """Определяет наличие по ячейке остатка/наличия."""
    if value is None:
        return False
    if isinstance(value, (int, float)):
        return value > 0
    s = str(value).strip().lower()
    if s in ('', 'нет', 'no', 'false', '0', '-', 'отсутствует'):
        return False
    if s in ('+', 'да', 'yes', 'true', 'в наличии'):
        return True
    n = parse_price(s)
    return bool(n)


class Command(BaseCommand):
    help = 'Полный импорт каталога из прайс-листов поставщиков'

    def handle(self, *args, **options):
        self.cat_cache = {}        # имя -> Category (верхнего уровня)
        self.products = {}         # имя товара -> Product (дедупликация)

        self.stdout.write('Очистка базы…')
        self._wipe()

        self.stdout.write('Импорт фид.yml …')
        self._import_yml(os.path.join(DOWNLOADS, 'фид.yml'))

        self._import_lamed(os.path.join(DOWNLOADS, 'lamed_kz_17379.xlsx'))
        self._import_ecogr(os.path.join(
            DOWNLOADS, 'Полный прайс на 06.05.26 Эко Групп Казахстан (1).xlsx'))
        self._import_alstyle(os.path.join(DOWNLOADS, 'Al-Style_price (1).xlsx'))

        self.stdout.write(f'Сохранение {len(self.products)} товаров…')
        with transaction.atomic():
            Product.objects.bulk_create(
                list(self.products.values()), batch_size=500)

        # Удаляем категории, в которых не осталось товаров.
        removed = 0
        while True:
            ids = list(Category.objects.filter(
                products__isnull=True, children__isnull=True
            ).values_list('pk', flat=True)[:400])
            if not ids:
                break
            Category.objects.filter(pk__in=ids).delete()
            removed += len(ids)

        self.stdout.write(self.style.SUCCESS(
            f'Готово. Товаров: {Product.objects.count()}, '
            f'категорий: {Category.objects.count()} '
            f'(удалено пустых: {removed}).'))

    # ─── очистка ──────────────────────────────────────────────────────────
    def _wipe(self):
        from django.db import connection
        OrderItem.objects.filter(product__isnull=False).update(product=None)
        # Прямой SQL вместо ORM-каскада — у Category есть ссылка на себя,
        # и сборщик каскадов упирается в лимит переменных SQLite.
        with connection.cursor() as cur:
            cur.execute('DELETE FROM shop_product')
            cur.execute('UPDATE shop_category SET parent_id = NULL')
            cur.execute('DELETE FROM shop_category')

    # ─── категории ────────────────────────────────────────────────────────
    def _make_category(self, name, parent=None):
        base = slugify(name, allow_unicode=True) or 'cat'
        slug, n = base, 2
        while Category.objects.filter(slug=slug).exists():
            slug = f'{base}-{n}'
            n += 1
        return Category.objects.create(name=name[:200], slug=slug, parent=parent)

    def get_category(self, name, parent=None):
        """Возвращает (создаёт при необходимости) категорию верхнего уровня
        либо подкатегорию указанного родителя."""
        key = (name, parent.pk if parent else None)
        if key in self.cat_cache:
            return self.cat_cache[key]
        cat = Category.objects.filter(
            name=name[:200], parent=parent).first() or \
            self._make_category(name, parent)
        self.cat_cache[key] = cat
        return cat

    def category_for(self, name):
        """Категория для товара-инструмента по названию (классификатор)."""
        cat_name = classify(name) or MISC_CATEGORY
        return self.get_category(cat_name)

    # ─── добавление товара с дедупликацией ────────────────────────────────
    def add_product(self, name, category, **fields):
        name = (name or '').strip()[:500]
        if not name or category is None:
            return
        existing = self.products.get(name)
        if existing is None:
            self.products[name] = Product(name=name, category=category, **fields)
        else:
            # дубль названия — обновляем поля более свежими данными
            existing.category = category
            for k, v in fields.items():
                setattr(existing, k, v)

    # ─── YML-фид ──────────────────────────────────────────────────────────
    def _import_yml(self, path):
        yml_cats, offers = {}, []
        for _ev, el in ET.iterparse(path, events=('end',)):
            if el.tag == 'category':
                cid = el.get('id')
                if cid:
                    yml_cats[cid] = {
                        'name': (el.text or '').strip() or 'Без названия',
                        'parent': el.get('parentId'),
                    }
                el.clear()
            elif el.tag == 'offer':
                row = self._yml_offer(el)
                if row:
                    offers.append(row)
                el.clear()

        resolved = {}

        def resolve(yid, depth=0):
            if yid in resolved:
                return resolved[yid]
            info = yml_cats.get(yid)
            if not info or depth > 12:
                return None
            parent = resolve(info['parent'], depth + 1) if info['parent'] else None
            cat = self.get_category(info['name'], parent)
            resolved[yid] = cat
            return cat

        count = 0
        for row in offers:
            cat = resolve(row.pop('cat_yml_id'))
            if cat is None:
                cat = self.get_category(MISC_CATEGORY)
            self.add_product(row.pop('name'), cat, **row)
            count += 1
        self.stdout.write(f'  фид.yml: {count} товаров, '
                          f'{len(yml_cats)} категорий')

    def _yml_offer(self, offer):
        def txt(tag):
            el = offer.find(tag)
            return (el.text or '').strip() if el is not None and el.text else ''

        name = txt('name')
        if not name:
            return None
        price = None
        for tag in ('price_rrc', 'price'):
            p = parse_price(txt(tag))
            if p:
                price = p
                break
        if price is None:
            return None

        pics = []
        for pic in offer.findall('picture'):
            u = (pic.text or '').strip()
            if u.lower().startswith(('http://', 'https://')) and u not in pics:
                pics.append(u)

        params = []
        for p in offer.findall('param'):
            pn = (p.get('name') or '').strip()
            pv = (p.text or '').strip()
            if pn:
                params.append(f'{pn}: {pv}' if pv else pn)

        src = txt('url')
        return dict(
            name=name[:500],
            cat_yml_id=txt('categoryId'),
            brand=txt('vendor')[:200],
            article=txt('vendorCode')[:200],
            price=price,
            image_url=pics[0] if pics else '',
            gallery='\n'.join(pics),
            source_url=src if src.lower().startswith(('http', 'https')) else '',
            description=strip_html(txt('description'))[:6000],
            characteristics='\n'.join(params)[:4000],
            in_stock=(offer.get('available') or 'true').lower() != 'false',
        )

    # ─── lamed.kz (классификация по названию) ─────────────────────────────
    def _import_lamed(self, path):
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        count = 0
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                continue
            name = str(row[1]).strip() if row[1] else ''
            price = parse_price(row[5])
            if not name or price is None:
                continue
            # характеристики из троек колонок 30..101
            chars = []
            for c in range(30, min(len(row), 102), 3):
                cn = str(row[c]).strip() if row[c] else ''
                cv = str(row[c + 2]).strip() if c + 2 < len(row) and row[c + 2] else ''
                cm = str(row[c + 1]).strip() if c + 1 < len(row) and row[c + 1] else ''
                if cn and cv:
                    chars.append(f'{cn}: {cv}{(" " + cm) if cm else ""}')
            # в ячейке со ссылкой на фото может быть несколько URL через запятую
            img_raw = str(row[11]).strip() if row[11] else ''
            pics = [u.strip() for u in img_raw.split(',')
                    if u.strip().lower().startswith(('http://', 'https://'))]
            self.add_product(
                name, self.category_for(name),
                brand=str(row[24]).strip()[:200] if row[24] else '',
                article=str(row[0]).strip()[:200] if row[0] else '',
                price=price,
                description=strip_html(row[3])[:6000],
                characteristics='\n'.join(chars)[:4000],
                image_url=pics[0] if pics else '',
                gallery='\n'.join(pics),
                in_stock=as_stock(row[12]) or as_stock(row[13]),
            )
            count += 1
        wb.close()
        self.stdout.write(f'  lamed.kz: {count} товаров')

    # ─── Эко Групп / ecogr.kz (классификация по названию) ─────────────────
    def _import_ecogr(self, path):
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        count = 0
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                continue
            name = str(row[1]).strip() if row[1] else ''
            price = parse_price(row[6]) or parse_price(row[5])
            if not name or price is None:
                continue
            src = str(row[4]).strip() if row[4] else ''
            self.add_product(
                name, self.category_for(name),
                brand=str(row[2]).strip()[:200] if row[2] else '',
                article=str(row[3]).strip()[:200] if row[3] else '',
                price=price,
                source_url=src if src.lower().startswith(('http', 'https')) else '',
                in_stock=as_stock(row[9]),
            )
            count += 1
        wb.close()
        self.stdout.write(f'  Эко Групп: {count} товаров')

    # ─── Al-Style (категории из разделов прайса) ──────────────────────────
    def _import_alstyle(self, path):
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        parent = self.get_category(ELECTRONICS_PARENT)
        section = None
        count = 0
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                continue
            code = str(row[0]).strip() if row[0] else ''
            name = str(row[2]).strip() if row[2] else ''
            # строка-раздел: есть текст в первой колонке, но нет названия товара
            if code and not name:
                section = self.get_category(code, parent)
                continue
            price = parse_price(row[5]) or parse_price(row[4])
            if not name or price is None:
                continue
            cat = section or self.get_category('Прочее', parent)
            self.add_product(
                name, cat,
                brand='',
                article=str(row[1]).strip()[:200] if row[1] else '',
                price=price,
                description=str(row[3]).strip()[:6000] if row[3] else '',
                in_stock=as_stock(row[7]),
            )
            count += 1
        wb.close()
        self.stdout.write(f'  Al-Style: {count} товаров')
